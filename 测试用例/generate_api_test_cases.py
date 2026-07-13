# -*- coding: utf-8 -*-
"""
云集优选(YunShop)接口测试用例生成器
生成Excel格式的接口测试用例
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime


def create_api_test_cases():
    """创建接口测试用例"""
    wb = openpyxl.Workbook()

    # ========== 样式定义 ==========
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    normal_font = Font(name='微软雅黑', size=10)
    normal_alignment = Alignment(vertical='center', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 优先级颜色
    priority_colors = {
        'P0': PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'),
        'P1': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),
        'P2': PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid'),
        'P3': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    }

    # ========== 表头定义 ==========
    headers = [
        '用例编号', '所属模块', '接口名称', '请求方式', '接口路径',
        '前置条件', '请求参数', '预期响应码', '预期响应数据', '优先级',
        '用例类型', '测试要点', '备注'
    ]

    # ========== 测试用例数据 ==========
    test_cases = [
        # ========== 用户模块 ==========
        ['TC_API_USER_001', '用户模块', '获取验证码', 'GET', '/api/user/captcha',
         '无', '无', 200, '返回data.captcha字段，长度为4位', 'P1',
         '功能测试', '验证验证码正常生成返回', ''],

        ['TC_API_USER_002', '用户模块', '会员登录-成功', 'POST', '/api/user/login',
         '已获取验证码', '{"username":"13800138000","password":"Test@123","captcha":"dev"}', 200,
         'code=200, msg="登录成功", data包含userId/mobile/nickname', 'P0',
         '功能测试', '验证正确账号密码登录成功', '开发环境captcha用dev跳过'],

        ['TC_API_USER_003', '用户模块', '会员登录-密码错误', 'POST', '/api/user/login',
         '已获取验证码', '{"username":"13800138000","password":"wrong123","captcha":"dev"}', 500,
         'code=500, msg="用户名或密码错误"', 'P0',
         '异常测试', '验证密码错误时返回错误提示', ''],

        ['TC_API_USER_004', '用户模块', '会员登录-用户不存在', 'POST', '/api/user/login',
         '已获取验证码', '{"username":"19999999999","password":"Test@123","captcha":"dev"}', 500,
         'code=500, msg="用户名或密码错误"', 'P1',
         '异常测试', '验证不存在的用户登录失败', ''],

        ['TC_API_USER_005', '用户模块', '会员登录-用户名为空', 'POST', '/api/user/login',
         '已获取验证码', '{"username":"","password":"Test@123","captcha":"dev"}', 500,
         'code=500, msg包含"用户名不能为空"', 'P1',
         '异常测试', '验证用户名为空时的校验', ''],

        ['TC_API_USER_006', '用户模块', '会员登录-密码为空', 'POST', '/api/user/login',
         '已获取验证码', '{"username":"13800138000","password":"","captcha":"dev"}', 500,
         'code=500, msg包含"密码不能为空"', 'P1',
         '异常测试', '验证密码为空时的校验', ''],

        ['TC_API_USER_007', '用户模块', '会员登录-验证码错误', 'POST', '/api/user/login',
         '已获取验证码', '{"username":"13800138000","password":"Test@123","captcha":"XXXX"}', 500,
         'code=500, msg="验证码错误"', 'P1',
         '异常测试', '验证验证码错误时登录失败', ''],

        ['TC_API_USER_008', '用户模块', '会员注册-手机号成功', 'POST', '/api/user/register',
         '无', '{"account":"13900139000","registerType":"mobile","password":"Test@123","confirmPassword":"Test@123","agreeProtocol":true}', 200,
         'code=200, msg="注册成功"', 'P0',
         '功能测试', '验证手机号注册成功', ''],

        ['TC_API_USER_009', '用户模块', '会员注册-邮箱成功', 'POST', '/api/user/register',
         '无', '{"account":"test01@163.com","registerType":"email","password":"Test@123","confirmPassword":"Test@123","agreeProtocol":true}', 200,
         'code=200, msg="注册成功"', 'P1',
         '功能测试', '验证邮箱注册成功', ''],

        ['TC_API_USER_010', '用户模块', '会员注册-手机号已存在', 'POST', '/api/user/register',
         '手机号已注册', '{"account":"13800138000","registerType":"mobile","password":"Test@123","confirmPassword":"Test@123","agreeProtocol":true}', 500,
         'code=500, msg包含"已注册"', 'P1',
         '异常测试', '验证重复手机号注册失败', ''],

        ['TC_API_USER_011', '用户模块', '会员注册-密码不一致', 'POST', '/api/user/register',
         '无', '{"account":"13900139001","registerType":"mobile","password":"Test@123","confirmPassword":"Test@456","agreeProtocol":true}', 500,
         'code=500, msg包含"密码不一致"', 'P1',
         '异常测试', '验证两次密码不一致时注册失败', ''],

        ['TC_API_USER_012', '用户模块', '会员注册-未同意协议', 'POST', '/api/user/register',
         '无', '{"account":"13900139002","registerType":"mobile","password":"Test@123","confirmPassword":"Test@123","agreeProtocol":false}', 500,
         'code=500, msg包含"协议"', 'P1',
         '异常测试', '验证未同意协议时注册失败', ''],

        ['TC_API_USER_013', '用户模块', '会员注册-手机号格式错误', 'POST', '/api/user/register',
         '无', '{"account":"12345","registerType":"mobile","password":"Test@123","confirmPassword":"Test@123","agreeProtocol":true}', 500,
         'code=500, msg包含"手机号"', 'P2',
         '异常测试', '验证手机号格式校验', '11位，第一位1，第二位非2'],

        ['TC_API_USER_014', '用户模块', '会员注册-密码强度不足', 'POST', '/api/user/register',
         '无', '{"account":"13900139003","registerType":"mobile","password":"123456","confirmPassword":"123456","agreeProtocol":true}', 500,
         'code=500, msg包含"密码"', 'P2',
         '异常测试', '验证密码强度校验', '需大小写+数字+符号至少两种'],

        ['TC_API_USER_015', '用户模块', '获取当前用户-已登录', 'GET', '/api/user/current',
         '已登录', '无', 200, 'code=200, data包含userId/mobile/nickname', 'P0',
         '功能测试', '验证已登录用户获取信息成功', ''],

        ['TC_API_USER_016', '用户模块', '获取当前用户-未登录', 'GET', '/api/user/current',
         '未登录', '无', 401, 'code=401, msg="未登录"', 'P1',
         '异常测试', '验证未登录时获取用户信息失败', ''],

        ['TC_API_USER_017', '用户模块', '登录状态检查-已登录', 'GET', '/api/user/check',
         '已登录', '无', 200, 'code=200, data=true', 'P1',
         '功能测试', '验证已登录状态返回true', ''],

        ['TC_API_USER_018', '用户模块', '登录状态检查-未登录', 'GET', '/api/user/check',
         '未登录', '无', 200, 'code=200, data=false', 'P1',
         '功能测试', '验证未登录状态返回false', ''],

        ['TC_API_USER_019', '用户模块', '更新个人信息-成功', 'PUT', '/api/user/profile',
         '已登录', '{"nickname":"新昵称","avatar":"/upload/avatars/test.png"}', 200,
         'code=200, msg="更新成功"', 'P0',
         '功能测试', '验证更新昵称和头像成功', ''],

        ['TC_API_USER_020', '用户模块', '更新个人信息-未登录', 'PUT', '/api/user/profile',
         '未登录', '{"nickname":"新昵称"}', 401, 'code=401, msg="未登录"', 'P1',
         '异常测试', '验证未登录时更新信息失败', ''],

        ['TC_API_USER_021', '用户模块', '修改密码-成功', 'PUT', '/api/user/password',
         '已登录', '{"oldPassword":"Test@123","newPassword":"NewPass@456"}', 200,
         'code=200, msg="密码修改成功"', 'P0',
         '功能测试', '验证修改密码成功', ''],

        ['TC_API_USER_022', '用户模块', '修改密码-原密码错误', 'PUT', '/api/user/password',
         '已登录', '{"oldPassword":"wrong123","newPassword":"NewPass@456"}', 500,
         'code=500, msg包含"原密码"', 'P1',
         '异常测试', '验证原密码错误时修改失败', ''],

        ['TC_API_USER_023', '用户模块', '修改密码-未登录', 'PUT', '/api/user/password',
         '未登录', '{"oldPassword":"Test@123","newPassword":"NewPass@456"}', 401,
         'code=401, msg="未登录"', 'P1',
         '异常测试', '验证未登录时修改密码失败', ''],

        ['TC_API_USER_024', '用户模块', '安全退出', 'POST', '/api/user/logout',
         '已登录', '无', 200, 'code=200, msg="已安全退出"', 'P0',
         '功能测试', '验证退出登录成功', ''],

        # ========== 商品模块 ==========
        ['TC_API_GOODS_001', '商品模块', '商品列表-默认查询', 'GET', '/api/goods/list',
         '无', '无', 200, 'code=200, data包含total/page/pageSize/list', 'P0',
         '功能测试', '验证商品列表默认分页返回', ''],

        ['TC_API_GOODS_002', '商品模块', '商品列表-按分类筛选', 'GET', '/api/goods/list?catId=1',
         '无', 'catId=1', 200, '返回该分类下商品，list不为空', 'P1',
         '功能测试', '验证按分类筛选商品', ''],

        ['TC_API_GOODS_003', '商品模块', '商品列表-关键词搜索', 'GET', '/api/goods/list?keyword=华为',
         '无', 'keyword=华为', 200, '返回包含关键词的商品', 'P1',
         '功能测试', '验证关键词搜索商品', ''],

        ['TC_API_GOODS_004', '商品模块', '商品列表-价格升序', 'GET', '/api/goods/list?sort=price&order=asc',
         '无', 'sort=price&order=asc', 200, '商品按价格升序排列', 'P2',
         '功能测试', '验证价格升序排序', ''],

        ['TC_API_GOODS_005', '商品模块', '商品列表-价格降序', 'GET', '/api/goods/list?sort=price&order=desc',
         '无', 'sort=price&order=desc', 200, '商品按价格降序排列', 'P2',
         '功能测试', '验证价格降序排序', ''],

        ['TC_API_GOODS_006', '商品模块', '商品列表-销量排序', 'GET', '/api/goods/list?sort=sales&order=desc',
         '无', 'sort=sales&order=desc', 200, '商品按销量降序排列', 'P2',
         '功能测试', '验证销量排序', ''],

        ['TC_API_GOODS_007', '商品模块', '商品列表-分页', 'GET', '/api/goods/list?page=2&pageSize=12',
         '无', 'page=2&pageSize=12', 200, '返回第2页数据，page=2', 'P1',
         '功能测试', '验证分页功能', ''],

        ['TC_API_GOODS_008', '商品模块', '商品列表-每页条数', 'GET', '/api/goods/list?pageSize=24',
         '无', 'pageSize=24', 200, '返回最多24条数据', 'P2',
         '功能测试', '验证每页条数设置', ''],

        ['TC_API_GOODS_009', '商品模块', '商品详情-正常', 'GET', '/api/goods/detail/1',
         '商品ID=1存在且上架', '无', 200, 'data包含goods/images/attrs', 'P0',
         '功能测试', '验证获取商品详情成功', ''],

        ['TC_API_GOODS_010', '商品模块', '商品详情-不存在', 'GET', '/api/goods/detail/99999',
         '无', '无', 500, 'code=500, msg包含"不存在"', 'P1',
         '异常测试', '验证商品不存在时返回错误', ''],

        ['TC_API_GOODS_011', '商品模块', '商品详情-已下架', 'GET', '/api/goods/detail/{下架商品ID}',
         '商品已下架', '无', 500, 'code=500, msg包含"下架"', 'P1',
         '异常测试', '验证已下架商品返回错误', ''],

        ['TC_API_GOODS_012', '商品模块', '热门推荐', 'GET', '/api/goods/hot',
         '无', '无', 200, '返回推荐商品列表', 'P1',
         '功能测试', '验证热门推荐接口', ''],

        ['TC_API_GOODS_013', '商品模块', '热门推荐-指定数量', 'GET', '/api/goods/hot?limit=5',
         '无', 'limit=5', 200, '返回最多5条推荐商品', 'P2',
         '功能测试', '验证限制推荐数量', ''],

        ['TC_API_GOODS_014', '商品模块', '关键词搜索', 'GET', '/api/goods/search?keyword=手机',
         '无', 'keyword=手机', 200, '返回搜索结果，包含分页信息', 'P1',
         '功能测试', '验证搜索接口', ''],

        # ========== 购物车模块 ==========
        ['TC_API_CART_001', '购物车模块', '购物车列表', 'GET', '/api/cart/list',
         '已登录', '无', 200, '返回购物车商品列表', 'P0',
         '功能测试', '验证获取购物车列表', ''],

        ['TC_API_CART_002', '购物车模块', '购物车列表-未登录', 'GET', '/api/cart/list',
         '未登录', '无', 401, 'code=401或提示登录', 'P1',
         '异常测试', '验证未登录获取购物车失败', ''],

        ['TC_API_CART_003', '购物车模块', '加入购物车-成功', 'POST', '/api/cart/add',
         '已登录', '{"goodsId":1,"num":1,"attrId":0}', 200, 'code=200, msg="已加入购物车"', 'P0',
         '功能测试', '验证加入购物车成功', ''],

        ['TC_API_CART_004', '购物车模块', '加入购物车-未登录', 'POST', '/api/cart/add',
         '未登录', '{"goodsId":1,"num":1}', 401, '提示需要登录', 'P1',
         '异常测试', '验证未登录加入购物车失败', ''],

        ['TC_API_CART_005', '购物车模块', '加入购物车-累加数量', 'POST', '/api/cart/add',
         '已登录，商品已在购物车', '{"goodsId":1,"num":1}', 200, 'code=200，数量累加', 'P1',
         '功能测试', '验证重复添加时数量累加', ''],

        ['TC_API_CART_006', '购物车模块', '加入购物车-数量超限', 'POST', '/api/cart/add',
         '已登录', '{"goodsId":1,"num":201}', 500, 'code=500, 提示数量超限', 'P2',
         '异常测试', '验证单品最大200数量限制', ''],

        ['TC_API_CART_007', '购物车模块', '修改数量-成功', 'PUT', '/api/cart/update/1',
         '已登录，购物车ID=1存在', '{"num":3}', 200, 'code=200，数量修改成功', 'P0',
         '功能测试', '验证修改购物车数量', ''],

        ['TC_API_CART_008', '购物车模块', '修改数量-无效ID', 'PUT', '/api/cart/update/99999',
         '已登录', '{"num":3}', 500, 'code=500，提示不存在', 'P2',
         '异常测试', '验证无效购物车ID', ''],

        ['TC_API_CART_009', '购物车模块', '删除商品-成功', 'DELETE', '/api/cart/delete/1',
         '已登录，购物车ID=1存在', '无', 200, 'code=200，删除成功', 'P0',
         '功能测试', '验证删除购物车商品', ''],

        ['TC_API_CART_010', '购物车模块', '批量删除-成功', 'POST', '/api/cart/delete-batch',
         '已登录', '{"ids":[1,2,3]}', 200, 'code=200，批量删除成功', 'P1',
         '功能测试', '验证批量删除功能', ''],

        ['TC_API_CART_011', '购物车模块', '切换选中状态', 'PUT', '/api/cart/toggle/1',
         '已登录，购物车ID=1存在', '无', 200, 'code=200，状态切换成功', 'P1',
         '功能测试', '验证切换选中状态', ''],

        ['TC_API_CART_012', '购物车模块', '全选', 'PUT', '/api/cart/select-all',
         '已登录', '{"selected":true}', 200, 'code=200，全选成功', 'P1',
         '功能测试', '验证全选功能', ''],

        ['TC_API_CART_013', '购物车模块', '取消全选', 'PUT', '/api/cart/select-all',
         '已登录', '{"selected":false}', 200, 'code=200，取消全选成功', 'P1',
         '功能测试', '验证取消全选功能', ''],

        ['TC_API_CART_014', '购物车模块', '购物车数量-已登录', 'GET', '/api/cart/count',
         '已登录', '无', 200, 'code=200, data为商品种类数', 'P1',
         '功能测试', '验证获取购物车数量', ''],

        ['TC_API_CART_015', '购物车模块', '购物车数量-未登录', 'GET', '/api/cart/count',
         '未登录', '无', 200, 'code=200, data=0', 'P2',
         '功能测试', '验证未登录时购物车数量为0', ''],

        # ========== 订单模块 ==========
        ['TC_API_ORDER_001', '订单模块', '提交订单-成功', 'POST', '/api/order/submit',
         '已登录，有购物车商品，有收货地址', '{"addressId":1,"cartIds":"1,2","payName":"alipay","shippingName":"express","remark":"尽快发货"}', 200,
         'code=200, msg="下单成功", data包含orderId/orderSn', 'P0',
         '功能测试', '验证提交订单成功', ''],

        ['TC_API_ORDER_002', '订单模块', '提交订单-未登录', 'POST', '/api/order/submit',
         '未登录', '{"addressId":1,"cartIds":"1"}', 401, '提示需要登录', 'P1',
         '异常测试', '验证未登录提交订单失败', ''],

        ['TC_API_ORDER_003', '订单模块', '提交订单-地址为空', 'POST', '/api/order/submit',
         '已登录', '{"cartIds":"1"}', 500, 'code=500，提示地址必填', 'P2',
         '异常测试', '验证缺少地址时提交失败', ''],

        ['TC_API_ORDER_004', '订单模块', '提交订单-购物车为空', 'POST', '/api/order/submit',
         '已登录，购物车为空', '{"addressId":1,"cartIds":""}', 500, 'code=500，提示购物车为空', 'P2',
         '异常测试', '验证购物车为空时提交失败', ''],

        ['TC_API_ORDER_005', '订单模块', '订单列表', 'GET', '/api/order/list',
         '已登录', '无', 200, '返回订单列表', 'P0',
         '功能测试', '验证获取订单列表', ''],

        ['TC_API_ORDER_006', '订单模块', '订单列表-按状态筛选', 'GET', '/api/order/list?status=PENDING',
         '已登录', 'status=PENDING', 200, '返回待付款订单', 'P1',
         '功能测试', '验证按状态筛选订单', ''],

        ['TC_API_ORDER_007', '订单模块', '订单列表-未登录', 'GET', '/api/order/list',
         '未登录', '无', 401, '提示需要登录', 'P1',
         '异常测试', '验证未登录获取订单列表失败', ''],

        ['TC_API_ORDER_008', '订单模块', '订单详情-成功', 'GET', '/api/order/detail/1',
         '已登录，订单属于当前用户', '无', 200, 'data包含order和goodsList', 'P0',
         '功能测试', '验证获取订单详情', ''],

        ['TC_API_ORDER_009', '订单模块', '订单详情-不存在', 'GET', '/api/order/detail/99999',
         '已登录', '无', 500, 'code=500, msg="订单不存在"', 'P1',
         '异常测试', '验证订单不存在时返回错误', ''],

        ['TC_API_ORDER_010', '订单模块', '订单详情-非本人订单', 'GET', '/api/order/detail/{他人订单ID}',
         '已登录', '无', 500, 'code=500, msg="订单不存在"', 'P2',
         '异常测试', '验证无法查看他人订单', ''],

        ['TC_API_ORDER_011', '订单模块', '模拟支付-成功', 'PUT', '/api/order/pay/1',
         '已登录，订单状态为PENDING', '无', 200, 'code=200, msg="支付成功"', 'P0',
         '功能测试', '验证模拟支付成功', ''],

        ['TC_API_ORDER_012', '订单模块', '模拟支付-已支付订单', 'PUT', '/api/order/pay/{已支付订单ID}',
         '已登录，订单状态为PAID', '无', 500, 'code=500，提示状态错误', 'P2',
         '异常测试', '验证重复支付失败', ''],

        ['TC_API_ORDER_013', '订单模块', '取消订单-成功', 'PUT', '/api/order/cancel/1',
         '已登录，订单状态为PENDING', '无', 200, 'code=200, msg="订单已取消"', 'P0',
         '功能测试', '验证取消订单成功', ''],

        ['TC_API_ORDER_014', '订单模块', '取消订单-已付款订单', 'PUT', '/api/order/cancel/{已付款订单ID}',
         '已登录，订单状态为PAID', '无', 500, 'code=500，提示状态错误', 'P2',
         '异常测试', '验证已付款订单无法取消', ''],

        ['TC_API_ORDER_015', '订单模块', '确认收货-成功', 'PUT', '/api/order/receive/1',
         '已登录，订单状态为SHIPPED', '无', 200, 'code=200, msg="已确认收货"', 'P0',
         '功能测试', '验证确认收货成功', ''],

        ['TC_API_ORDER_016', '订单模块', '确认收货-未发货订单', 'PUT', '/api/order/receive/{未发货订单ID}',
         '已登录，订单状态为PENDING或PAID', '无', 500, 'code=500，提示状态错误', 'P2',
         '异常测试', '验证未发货订单无法确认收货', ''],

        # ========== 收货地址模块 ==========
        ['TC_API_ADDR_001', '地址模块', '地址列表', 'GET', '/api/address/list',
         '已登录', '无', 200, '返回地址列表', 'P0',
         '功能测试', '验证获取地址列表', ''],

        ['TC_API_ADDR_002', '地址模块', '地址列表-未登录', 'GET', '/api/address/list',
         '未登录', '无', 401, '提示需要登录', 'P1',
         '异常测试', '验证未登录获取地址失败', ''],

        ['TC_API_ADDR_003', '地址模块', '获取单个地址-成功', 'GET', '/api/address/1',
         '已登录，地址ID=1存在', '无', 200, '返回地址详情', 'P1',
         '功能测试', '验证获取单个地址', ''],

        ['TC_API_ADDR_004', '地址模块', '获取单个地址-不存在', 'GET', '/api/address/99999',
         '已登录', '无', 500, 'code=500, msg="地址不存在"', 'P2',
         '异常测试', '验证地址不存在时返回错误', ''],

        ['TC_API_ADDR_005', '地址模块', '新增地址-成功', 'POST', '/api/address/add',
         '已登录', '{"consignee":"张三","mobile":"13800138000","province":"广东省","city":"深圳市","district":"南山区","address":"科技园路100号"}', 200,
         'code=200, msg="添加成功"', 'P0',
         '功能测试', '验证新增地址成功', ''],

        ['TC_API_ADDR_006', '地址模块', '新增地址-未登录', 'POST', '/api/address/add',
         '未登录', '{"consignee":"张三","mobile":"13800138000","province":"广东省","city":"深圳市","district":"南山区","address":"科技园路100号"}', 401,
         '提示需要登录', 'P1',
         '异常测试', '验证未登录新增地址失败', ''],

        ['TC_API_ADDR_007', '地址模块', '新增地址-数量超限', 'POST', '/api/address/add',
         '已登录，已有20个地址', '{"consignee":"张三",...}', 500, 'code=500，提示地址数量超限', 'P2',
         '异常测试', '验证地址最多20个限制', ''],

        ['TC_API_ADDR_008', '地址模块', '编辑地址-成功', 'PUT', '/api/address/update',
         '已登录', '{"addressId":1,"consignee":"李四","mobile":"13900139000",...}', 200,
         'code=200, msg="更新成功"', 'P0',
         '功能测试', '验证编辑地址成功', ''],

        ['TC_API_ADDR_009', '地址模块', '删除地址-成功', 'DELETE', '/api/address/delete/1',
         '已登录，地址ID=1存在', '无', 200, 'code=200, msg="删除成功"', 'P0',
         '功能测试', '验证删除地址成功', ''],

        ['TC_API_ADDR_010', '地址模块', '删除地址-不存在', 'DELETE', '/api/address/delete/99999',
         '已登录', '无', 500, 'code=500，提示不存在', 'P2',
         '异常测试', '验证删除不存在地址', ''],

        ['TC_API_ADDR_011', '地址模块', '设为默认地址', 'PUT', '/api/address/default/1',
         '已登录，地址ID=1存在', '无', 200, 'code=200, msg="已设为默认地址"', 'P1',
         '功能测试', '验证设为默认地址', ''],

        # ========== 首页数据模块 ==========
        ['TC_API_HOME_001', '首页模块', '获取首页数据', 'GET', '/api/home',
         '无', '无', 200, 'data包含navList/categoryMenu/bannerList/floors/hotGoods', 'P0',
         '功能测试', '验证首页数据接口', ''],

        ['TC_API_HOME_002', '首页模块', '获取分类菜单', 'GET', '/api/categories/menu',
         '无', '无', 200, '返回三级分类树结构', 'P1',
         '功能测试', '验证分类菜单接口', ''],

        # ========== 图片接口 ==========
        ['TC_API_IMG_001', '图片接口', '商品图片-存在', 'GET', '/api/images/goods/1',
         '商品ID=1存在', '无', 200, '返回PNG图片，Content-Type为image', 'P2',
         '功能测试', '验证商品图片接口', ''],

        ['TC_API_IMG_002', '图片接口', '商品图片-不存在', 'GET', '/api/images/goods/99999',
         '无', '无', 200, '返回默认占位图', 'P3',
         '功能测试', '验证不存在商品返回占位图', ''],

        ['TC_API_IMG_003', '图片接口', 'Banner图片', 'GET', '/api/images/banner/1',
         '无', '无', 200, '返回PNG图片', 'P3',
         '功能测试', '验证Banner图片接口', ''],

        # ========== 文件上传 ==========
        ['TC_API_UPLOAD_001', '上传模块', '上传头像-成功', 'POST', '/api/upload/avatar',
         '已登录，准备PNG图片', 'multipart/form-data, file=avatar.png', 200,
         'code=200, data包含url和filename', 'P0',
         '功能测试', '验证上传头像成功', ''],

        ['TC_API_UPLOAD_002', '上传模块', '上传头像-未登录', 'POST', '/api/upload/avatar',
         '未登录', 'multipart/form-data, file=avatar.png', 401,
         '提示需要登录', 'P1',
         '异常测试', '验证未登录上传失败', ''],

        ['TC_API_UPLOAD_003', '上传模块', '上传头像-格式错误', 'POST', '/api/upload/avatar',
         '已登录', 'multipart/form-data, file=test.txt', 500,
         'code=500，提示格式不支持', 'P2',
         '异常测试', '验证不支持的文件格式', ''],

        ['TC_API_UPLOAD_004', '上传模块', '上传头像-文件过大', 'POST', '/api/upload/avatar',
         '已登录', 'multipart/form-data, file=超过5MB的图片', 500,
         'code=500，提示文件过大', 'P2',
         '异常测试', '验证文件大小限制', ''],

        # ========== 后台管理模块 ==========
        ['TC_API_ADMIN_001', '后台模块', '管理员登录-成功', 'POST', '/api/admin/login',
         '无', '{"username":"admin","password":"admin123"}', 200,
         'code=200, msg="登录成功", data包含adminId/username', 'P0',
         '功能测试', '验证管理员登录成功', ''],

        ['TC_API_ADMIN_002', '后台模块', '管理员登录-密码错误', 'POST', '/api/admin/login',
         '无', '{"username":"admin","password":"wrong123"}', 500,
         'code=500, msg="用户名或密码错误"', 'P1',
         '异常测试', '验证管理员密码错误', ''],

        ['TC_API_ADMIN_003', '后台模块', '管理员登录-用户不存在', 'POST', '/api/admin/login',
         '无', '{"username":"nonexist","password":"admin123"}', 500,
         'code=500, msg="用户名或密码错误"', 'P1',
         '异常测试', '验证不存在的管理员', ''],

        ['TC_API_ADMIN_004', '后台模块', '获取当前管理员', 'GET', '/api/admin/current',
         '管理员已登录', '无', 200, 'code=200, data包含管理员信息', 'P1',
         '功能测试', '验证获取管理员信息', ''],

        ['TC_API_ADMIN_005', '后台模块', '获取当前管理员-未登录', 'GET', '/api/admin/current',
         '未登录', '无', 401, 'code=401, msg="未登录"', 'P1',
         '异常测试', '验证未登录获取管理员信息', ''],

        ['TC_API_ADMIN_006', '后台模块', '管理员退出', 'POST', '/api/admin/logout',
         '管理员已登录', '无', 200, 'code=200, msg="已退出"', 'P1',
         '功能测试', '验证管理员退出', ''],

        ['TC_API_ADMIN_007', '后台模块', '添加商品-成功', 'POST', '/api/admin/goods/add',
         '管理员已登录', '{"goodsName":"测试商品","goodsSn":"SN20260713001","catId":1,"shopPrice":99.00,"storeCount":100,"isOnSale":1}', 200,
         'code=200，添加成功', 'P0',
         '功能测试', '验证添加商品成功', ''],

        ['TC_API_ADMIN_008', '后台模块', '添加商品-未登录', 'POST', '/api/admin/goods/add',
         '未登录', '{"goodsName":"测试商品",...}', 401, '提示需要登录', 'P1',
         '异常测试', '验证未登录添加商品失败', ''],

        ['TC_API_ADMIN_009', '后台模块', '更新商品', 'PUT', '/api/admin/goods/update',
         '管理员已登录', '{"goodsId":1,"goodsName":"更新商品","shopPrice":199.00}', 200,
         'code=200，更新成功', 'P0',
         '功能测试', '验证更新商品成功', ''],

        ['TC_API_ADMIN_010', '后台模块', '删除商品', 'DELETE', '/api/admin/goods/delete/1',
         '管理员已登录', '无', 200, 'code=200，删除成功（软删除）', 'P1',
         '功能测试', '验证删除商品成功', ''],

        ['TC_API_ADMIN_011', '后台模块', '上下架切换', 'PUT', '/api/admin/goods/toggle/1',
         '管理员已登录', '无', 200, 'code=200，状态切换成功', 'P1',
         '功能测试', '验证上下架切换', ''],
    ]

    # ========== 创建前台接口测试Sheet ==========
    ws_front = wb.active
    ws_front.title = '前台接口测试'

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws_front.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入前台测试数据（排除后台模块）
    row = 2
    for case in test_cases:
        if case[1] != '后台模块':
            for col, value in enumerate(case, 1):
                cell = ws_front.cell(row=row, column=col, value=value)
                cell.font = normal_font
                cell.alignment = normal_alignment if col in [6, 7, 9, 12] else center_alignment
                cell.border = thin_border
                # 优先级颜色
                if col == 10 and value in priority_colors:
                    cell.fill = priority_colors[value]
            row += 1

    # ========== 创建后台接口测试Sheet ==========
    ws_admin = wb.create_sheet('后台接口测试')

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws_admin.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入后台测试数据
    row = 2
    for case in test_cases:
        if case[1] == '后台模块':
            for col, value in enumerate(case, 1):
                cell = ws_admin.cell(row=row, column=col, value=value)
                cell.font = normal_font
                cell.alignment = normal_alignment if col in [6, 7, 9, 12] else center_alignment
                cell.border = thin_border
                # 优先级颜色
                if col == 10 and value in priority_colors:
                    cell.fill = priority_colors[value]
            row += 1

    # ========== 创建统计Sheet ==========
    ws_stats = wb.create_sheet('用例统计')

    # 统计数据
    total = len(test_cases)
    p0_count = sum(1 for c in test_cases if c[9] == 'P0')
    p1_count = sum(1 for c in test_cases if c[9] == 'P1')
    p2_count = sum(1 for c in test_cases if c[9] == 'P2')
    p3_count = sum(1 for c in test_cases if c[9] == 'P3')

    func_count = sum(1 for c in test_cases if c[10] == '功能测试')
    error_count = sum(1 for c in test_cases if c[10] == '异常测试')

    front_count = sum(1 for c in test_cases if c[1] != '后台模块')
    admin_count = sum(1 for c in test_cases if c[1] == '后台模块')

    stats_data = [
        ['云集优选接口测试用例统计', '', '', ''],
        ['', '', '', ''],
        ['统计维度', '分类', '数量', '占比'],
        ['总计', '全部用例', total, '100%'],
        ['', '', '', ''],
        ['按优先级', 'P0-核心功能', p0_count, f'{p0_count/total*100:.1f}%'],
        ['', 'P1-重要功能', p1_count, f'{p1_count/total*100:.1f}%'],
        ['', 'P2-一般功能', p2_count, f'{p2_count/total*100:.1f}%'],
        ['', 'P3-边缘功能', p3_count, f'{p3_count/total*100:.1f}%'],
        ['', '', '', ''],
        ['按用例类型', '功能测试', func_count, f'{func_count/total*100:.1f}%'],
        ['', '异常测试', error_count, f'{error_count/total*100:.1f}%'],
        ['', '', '', ''],
        ['按接口位置', '前台接口', front_count, f'{front_count/total*100:.1f}%'],
        ['', '后台接口', admin_count, f'{admin_count/total*100:.1f}%'],
        ['', '', '', ''],
        ['按模块', '用户模块', sum(1 for c in test_cases if c[1] == '用户模块'), ''],
        ['', '商品模块', sum(1 for c in test_cases if c[1] == '商品模块'), ''],
        ['', '购物车模块', sum(1 for c in test_cases if c[1] == '购物车模块'), ''],
        ['', '订单模块', sum(1 for c in test_cases if c[1] == '订单模块'), ''],
        ['', '地址模块', sum(1 for c in test_cases if c[1] == '地址模块'), ''],
        ['', '首页模块', sum(1 for c in test_cases if c[1] == '首页模块'), ''],
        ['', '图片接口', sum(1 for c in test_cases if c[1] == '图片接口'), ''],
        ['', '上传模块', sum(1 for c in test_cases if c[1] == '上传模块'), ''],
        ['', '后台模块', sum(1 for c in test_cases if c[1] == '后台模块'), ''],
        ['', '', '', ''],
        ['生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), '', ''],
    ]

    for r, row_data in enumerate(stats_data, 1):
        for c, value in enumerate(row_data, 1):
            cell = ws_stats.cell(row=r, column=c, value=value)
            cell.font = Font(name='微软雅黑', size=11, bold=(r <= 1 or r == 3))
            cell.alignment = Alignment(vertical='center')
            if r == 1:
                cell.font = Font(name='微软雅黑', size=14, bold=True, color='4472C4')

    # ========== 设置列宽 ==========
    column_widths = {
        'A': 18, 'B': 12, 'C': 20, 'D': 10, 'E': 25,
        'F': 25, 'G': 45, 'H': 12, 'I': 35, 'J': 8,
        'K': 10, 'L': 25, 'M': 20
    }

    for ws in [ws_front, ws_admin]:
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        # 冻结首行
        ws.freeze_panes = 'A2'
        # 设置自动筛选
        ws.auto_filter.ref = f'A1:M{ws.max_row}'

    # 统计表列宽
    ws_stats.column_dimensions['A'].width = 15
    ws_stats.column_dimensions['B'].width = 20
    ws_stats.column_dimensions['C'].width = 10
    ws_stats.column_dimensions['D'].width = 10

    # ========== 保存文件 ==========
    output_path = 'G:/TestProject/测试用例/接口测试用例.xlsx'
    wb.save(output_path)
    print(f'[OK] 测试用例已生成: {output_path}')
    print(f'[STATS] 总计: {total} 条用例')
    print(f'   - P0: {p0_count} 条, P1: {p1_count} 条, P2: {p2_count} 条, P3: {p3_count} 条')
    print(f'   - 功能测试: {func_count} 条, 异常测试: {error_count} 条')
    print(f'   - 前台接口: {front_count} 条, 后台接口: {admin_count} 条')


if __name__ == '__main__':
    create_api_test_cases()
